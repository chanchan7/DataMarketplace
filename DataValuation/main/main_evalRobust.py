"""
evaluate the robustness (unbalanced data, adversarial, malicious, noisy label) of the data valuation method
"""
import sys

import torch

sys.path.append("..")
import numpy as np
import copy
import logging
import os
from openpyxl import Workbook, load_workbook
from sklearn.model_selection import train_test_split

# os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
# os.environ['CUDA_VISIBLE_DEVICES'] = '0'
from tqdm import tqdm

from dataloader import construct_dataowner, sample_from_dataowner
from proxy_models import cnn_data_to_acc, logistic_data_to_acc
from utils import model_name, collect_from_loader, seed_everything
from preprocess import noise, RandomizedResopnse, attack_pgd
from experimental_setting import experimental_setting
from train_uf import train_accuracy_line
from sample_utility import sample_utility, sample_utility_unbalance_owner
from train_adda import train_adda_uf
from train_uf import train_uf
from train_uf_y import train_uf_y
from train_src_net import pretraining_from_data, train_source_from_loader
from models.models import get_model, model_for_cifar10, model_for_mnist
from shapley import get_shapley

logger = logging.getLogger(__name__)


def main():
    logging.basicConfig(
        format='[%(asctime)s] - %(message)s',
        datefmt='%Y/%m/%d %H:%M:%S',
        level=logging.INFO,
        filename=logfile
    )
    ext_n, ds_n = model_name(args, src, tgt)
    ext_file = os.path.join(args.out_dir, ext_n)
    ds_file = os.path.join(args.out_dir, ds_n)

    rng = np.random.default_rng(args.seed)
    data_owner_X, data_owner_y, valid_data = construct_dataowner(
        args.data_dir, args.n_owners, args.valid_size, rng, tgt, args.data_size, args.unbalance)  # Unbalanced data owners
    
    ################
    ## Noisy lable##
    ################
    for i in range(args.n_owners):
        data_owner_y_noise = copy.copy(data_owner_y)
        if args.noisey and args.prob_hold is not None:
            mechanism = RandomizedResopnse(p=args.prob_hold, d=10)
            y_perturbed = mechanism(data_owner_y[i])
            data_owner_y_noise[i] = y_perturbed

    ######################
    ## Adversarial owner##
    ######################
    if args.adv_owners is not None:
        if tgt == 'cifar10':
            model_for_adv = model_for_cifar10().cuda()
        else:
            model_for_adv = model_for_mnist().cuda()
        state_dict = torch.load('../model_{}.pth'.format(tgt))
        model_for_adv.load_state_dict(state_dict)
        model_for_adv.eval()
        target_ind = [0]  # rng.choice(args.n_owners, args.adv_owners, replace=False)
        for ind in target_ind:
            data_owner_X[ind] = attack_pgd(model_for_adv, data_owner_X[ind], data_owner_y[ind], epsilon=args.epsilon,
                                           attack_iters=20, restarts=1)
        sample = rng.choice(args.valid_size, int(args.proportion * args.valid_size), replace=False)
        valid_data[0][sample] = attack_pgd(model_for_adv, valid_data[0][sample], valid_data[1][sample],
                                           epsilon=args.epsilon, attack_iters=20, restarts=1)

    kwargs = {'share': args.share, 'public_size': args.public_size, 'batch_size': args.batch_size}
    sample_loaders, tgt_index_range = sample_from_dataowner(
        data_owner_X, data_owner_y_noise, args.data_dir, args.n_owners, rng, src, kwargs)
    utility_data = collect_from_loader(sample_loaders)

    if os.path.exists(pretrain_net_file):
        print('Skipping pre-training net training, exists:', pretrain_net_file)
    else:
        pretraining_from_data(utility_data, args.model, args.num_cls, outfile=pretrain_net_file, num_epoch=100, batch=128, lr=1e-4)

    ####################
    ## Malicious owner##
    ####################
    if args.malicious_owners is not None:
        target_ind = rng.choice(args.n_owners, args.malicious_owners, replace=False)
        data_owner_X = noise(data_owner_X, target_ind=target_ind, scale=10)

    ####################
    ## Replicate owner##
    ####################
    if args.replicate is not None:
        target_ind = [0]
        for ind in target_ind:
            data_owner_X[ind] = data_owner_X[ind].repeat_interleave(args.replicate, 0)  # tensor
            data_owner_y[ind] = data_owner_y[ind].repeat(args.replicate)  # array

    if os.path.exists(ext_file) and os.path.exists(ds_file):
        print('Skipping extractor training and deepset training, '
              'exists extractor: {}, deepset: {}'.format(ext_file, ds_file))
        extractor = get_model(args.model, num_cls=args.num_cls, weights_init=pretrain_net_file).to(args.device)
        extractor.load(ext_file)

        hidden_num = 512
        deepset = get_model('DeepSets', in_features=extractor.output_dim, hidden_ext=hidden_num,
                            hidden_reg=hidden_num).to(args.device)
        deepset.load(ds_file)
        logger.info('Load Extractor proxy_model and DeepSet proxy_model for computing shapley value')

    else:
        ################################### generate utility function dataset ###################################
        if os.path.exists(x_path) and os.path.exists(y_path):
            X_feature = np.load(x_path, allow_pickle=True).tolist()
            y_feature = np.load(y_path).tolist()
        else:
            if src is None:
                set_size = utility_data[1].shape[0]
                X_feature, y_feature = sample_utility_unbalance_owner(4000, set_size // 2, set_size, proxy_model,
                                                                      (utility_data, valid_data), tgt_index_range,
                                                                      args.utility_seed, verbose=True)
            else:
                X_feature, y_feature = sample_utility(4000, 200, args.public_size, proxy_model,
                                                      (utility_data, valid_data),
                                                      args.utility_seed, verbose=True)
            np.save(x_path, np.array(X_feature, dtype=object))
            np.save(y_path, y_feature)

        X_feature, X_feature_test, y_feature, y_feature_test = train_test_split(X_feature, y_feature, test_size=0.2)
        y_feature = torch.tensor(y_feature, requires_grad=False, dtype=torch.float32)
        y_feature_test = torch.tensor(y_feature_test, requires_grad=False, dtype=torch.float32)

        ########################################### train utility function  ###########################################
        train_set, test_set = (X_feature, y_feature), (X_feature_test, y_feature_test)

        if src is not None:
            train_set, test_set = (X_feature, y_feature), (X_feature_test, y_feature_test)
            # tgt_sample = collect_from_loader(tgt_sample_loader)
            lr_adda = {'src': args.lr_ext_src, 'tgt': args.lr_ext_tgt, 'dis': args.lr_ext_dis}
            extractor, deepset = train_adda_uf(src, tgt, sample_loaders, utility_data[0], train_set, test_set,
                                               args.epochs, pretrain_net_file=pretrain_net_file, out_dir=args.out_dir,
                                               logger=logger, model=args.model, ext_file=ext_file, ds_file=ds_file,
                                               lr_adda=lr_adda, lr_deepsets=args.lr_deepsets)
        elif args.model == 'LDTN':
            extractor, deepset = train_uf_y(tgt, utility_data[0], utility_data[1], train_set, test_set, args.epochs,
                                            src_weights=pretrain_net_file,
                                            out_dir=args.out_dir, logger=logger,
                                            betas=args.betas, weight_decay=args.weight_decay,
                                            model=args.model,
                                            ext_file=ext_file, ds_file=ds_file)
        else:
            extractor, deepset = train_uf(tgt, utility_data[0], train_set, test_set, args.epochs,
                                          pretrain_net_file=pretrain_net_file, out_dir=args.out_dir, logger=logger,
                                          model=args.model, ext_file=ext_file, ds_file=ds_file,
                                          lr_extractor=args.lr_extractor, lr_deepsets=args.lr_deepsets)
        extractor.eval()
        deepset.eval()

    ########################################### calculate Shapley Values  ###########################################
    shapley_value_uf = get_shapley(args.n_owners, data_owner_X, data_owner_y, extractor, deepset)

    return shapley_value_uf, target_ind


if __name__ == '__main__':
    args = experimental_setting()
    seed_everything(args.seed)
    src = args.src
    tgt = args.tgt
    args.model = 'MPCLeNet' if tgt == 'mnist' else 'MPCDTN'
    args.num_cls = 10
    proxy_model = logistic_data_to_acc

    if args.unbalance:
        flag = 'unbalance'
    elif args.adv_owners is not None:
        flag = 'adv'
    elif args.malicious_owners is not None:
        flag = 'malicious'
    elif args.noisey:
        flag = 'noisey'
    elif args.replicate is not None:
        flag = 'replcate'


    if src is None:
        args.out_dir = os.path.join(args.out_dir, '{}/models/{}/'.format(flag, tgt))
        args.out_feature_dir = os.path.join(args.out_feature_dir, '{}/feature/{}'.format(flag, tgt))
        pretrain_net_file = os.path.join(args.out_dir, '{}_pretrain_{:s}.pth'.format(args.model, tgt))
        logfile = os.path.join(args.out_dir, 'output_{}_{}:{}.log'.format(tgt, args.n_owners, args.data_size))
        x_path = '{}/s1_x_{}:{}_{}.npy'.format(args.out_feature_dir, args.n_owners, args.data_size, tgt)
        y_path = '{}/s1_y_{}:{}_{}.npy'.format(args.out_feature_dir, args.n_owners, args.data_size, tgt)
        file_name = '../outputs/{}/models/{}/output_{}:{}.xlsx'.format(flag, tgt, args.n_owners, args.data_size)
        sheet_name = 's1_{}'.format(tgt)

    else:
        args.out_dir = os.path.join(args.out_dir, '{}/models/{}2{}/'.format(flag, src, tgt))
        args.out_feature_dir = os.path.join(args.out_feature_dir, '{}/feature/{}2{}'.format(flag, src, tgt))
        pretrain_net_file = os.path.join(args.out_dir, '{}_pretrain_{:s}.pth'.format(args.model, src))
        logfile = os.path.join(args.out_dir, 'output_{}2{}_{}:{}.log'.format(src, tgt, args.n_owners, args.data_size))
        x_path = '{}/s3_x_{}:{}_{}.npy'.format(args.out_feature_dir, args.n_owners, args.data_size, src)
        y_path = '{}/s3_y_{}:{}_{}.npy'.format(args.out_feature_dir, args.n_owners, args.data_size, src)
        file_name = '../outputs/{}/models/{}2{}/output_{}:{}.xlsx'.format(flag, src, tgt, args.n_owners, args.data_size)
        sheet_name = 's3_{}2{}'.format(src, tgt)

    if not os.path.exists(args.out_dir):
        os.makedirs(args.out_dir)
    if not os.path.exists(args.out_feature_dir):
        os.makedirs(args.out_feature_dir)
    if os.path.exists(logfile):
        os.remove(logfile)

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    ws = wb.create_sheet(sheet_name)
    progbar = tqdm(range(args.repeats), file=sys.stdout, position=0, leave=False)
    for i in progbar:
        args.seed = i
        args.utility_seed = i + 1
        shapley_value, target_index = main()
        row = [args.n_owners, args.data_size] + list(shapley_value) + list(target_index)
        ws.append(row)

    # # args = get_args()
    # # proxy_model = logistic_data_to_acc
    # # file_name = '../outputs/3/models/scenario1/Experiment3_scenario1.xlsx'
    # # # file_name_out = '../outputs/2_b/datasize_2b_metric.xlsx'
    # # wb = load_workbook(file_name)
    # # ws = wb.create_sheet('scenario1_mnist_malicious')
    # # # ws_metric = wb.create_sheet('scenario1_mnist_metric')
    # for prop in [0.2]:
    #     # for owner in [1, 2, 3, 4, 5, 6, 7]:
    #     args.proportion = prop
    #     # args.unbalance = True
    #     shapley_value, target_index = main('mnist', args)
    #     # acu = np.array([123,124,4,34,234,124]).tolist()
    #
    #     # row1 = [args.n_owners, owner] + list(shapley_value) + list(target_index)
    #     # ws.append(row1)
    #     # # ws.cell(row=row, column=3, value=test_auc)
    #     # wb.save(file_name)
