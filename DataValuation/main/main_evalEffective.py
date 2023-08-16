"""
Given the #/ data owners, vary #/ data points per owner
Given #/ data points per owner, vary #/ data points
"""
import sys

import torch

sys.path.append("..")
import numpy as np
import copy
import logging
import os
from openpyxl import Workbook, load_workbook
from sklearn.model_selection import train_test_split
from csv import writer

# os.environ["CUDA_DEVICE_ORDER"] = "PCI_BUS_ID"
# os.environ['CUDA_VISIBLE_DEVICES'] = '0'
from tqdm import tqdm

from dataloader import construct_dataowner, sample_from_dataowner
from proxy_models import cnn_data_to_acc, logistic_data_to_acc
from utils import model_name, collect_from_loader, seed_everything
from preprocess import addGaussianNoise, RandomizedResopnse
from experimental_setting import experimental_setting
from train_uf import train_accuracy_line
from sample_utility import sample_utility, sample_utility_unbalance_owner
from train_adda import train_adda_uf
from train_uf import train_uf
from train_uf_y import train_uf_y
from train_src_net import pretraining_from_data, train_source_from_loader
from models.models import get_model
from shapley import get_shapley

logger = logging.getLogger(__name__)


def main():
    logging.basicConfig(
        format='[%(asctime)s] - %(message)s',
        datefmt='%Y/%m/%d %H:%M:%S',
        level=logging.INFO,
        # filename=logfile,
    )
    logger.info(args)

    ext_n, ds_n = model_name(args, src, tgt)
    ext_file = os.path.join(args.out_dir, ext_n)
    ds_file = os.path.join(args.out_dir, ds_n)

    rng = np.random.default_rng(12345)
    data_owner_X, data_owner_y, valid_data = construct_dataowner(
        args.data_dir, args.n_owners, args.valid_size, rng, tgt, args.data_size, args.unbalance)

    for i in range(args.n_owners):
        scale = i * 1/args.n_owners  #rng.uniform(1, 10)
        data_owner_X[i] = addGaussianNoise(data_owner_X[i], scale=scale)
        data_owner_y_noise = copy.copy(data_owner_y)
        if args.noisey and args.prob_hold is not None:
            mechanism = RandomizedResopnse(p=args.prob_hold, d=10)
            y_perturbed = mechanism(data_owner_y[i])
            data_owner_y_noise[i] = y_perturbed

    kwargs = {'share': args.share, 'public_size': args.public_size, 'batch_size': args.batch_size}

    sample_loaders, tgt_index_range = sample_from_dataowner(
        data_owner_X, data_owner_y_noise, args.data_dir, args.n_owners, rng, src, kwargs)

    utility_data = collect_from_loader(sample_loaders)

    if os.path.exists(pretrain_net_file):
        print('Skipping pre-training net training, exists:', pretrain_net_file)
    else:
        pretraining_from_data(utility_data, args.model, args.num_cls, pretrain_net_file,
                              num_epoch=args.pre_train_epoch, batch=32, lr=1e-4)

    if os.path.exists(ext_file) and os.path.exists(ds_file):
        print('Skipping extractor training and deepset training, '
              'exists extractor: {}, deepset: {}'.format(ext_file, ds_file))
        extractor = get_model(args.model, num_cls=args.num_cls, weights_init=pretrain_net_file).to(args.device)
        extractor.load(ext_file)

        hidden_num = 512
        deepset = get_model('DeepSets', in_features=extractor.output_dim, hidden_ext=hidden_num,
                            hidden_reg=hidden_num).to(args.device)
        deepset.load(ds_file)
        logger.info('Load Extractor and DeepSet for computing shapley value')

    else:
        ################################### generate utility function dataset ###################################
        if os.path.exists(x_path) and os.path.exists(y_path):
            X_feature = np.load(x_path, allow_pickle=True).tolist()
            y_feature = np.load(y_path).tolist()
        else:
            if src is None:
                # set_size = utility_data[1].shape[0]
                X_feature, y_feature = sample_utility(6000, args.min_set, args.max_set, proxy_model,
                                                      (utility_data, valid_data),
                                                      args.utility_seed, verbose=True)
            else:
                X_feature, y_feature = sample_utility(4000, 20, 100, proxy_model,
                                                      (utility_data, valid_data),
                                                      args.utility_seed, verbose=True)
            np.save(x_path, np.array(X_feature, dtype=object))
            np.save(y_path, y_feature)

        X_feature, X_feature_test, y_feature, y_feature_test = train_test_split(X_feature, y_feature, test_size=0.2)
        y_feature = torch.tensor(y_feature, requires_grad=False, dtype=torch.float32)
        y_feature_test = torch.tensor(y_feature_test, requires_grad=False, dtype=torch.float32)

        ########################################### train utility function  ###########################################
        train_set, test_set = (X_feature, y_feature), (X_feature_test, y_feature_test)

        if src is not None:
            train_set, test_set = (X_feature, y_feature), (X_feature_test, y_feature_test)
            # tgt_sample = collect_from_loader(tgt_sample_loader)
            lr_adda = {'src': args.lr_ext_src, 'tgt': args.lr_ext_tgt, 'dis': args.lr_ext_dis}
            extractor, deepset = train_adda_uf(src, tgt, sample_loaders, utility_data[0], train_set, test_set,
                                               args.num_cls,
                                               args.epochs, pretrain_net_file=pretrain_net_file, out_dir=args.out_dir,
                                               logger=logger, model=args.model, ext_file=ext_file, ds_file=ds_file,
                                               lr_adda=lr_adda, lr_deepsets=args.lr_deepsets, device=args.device)
        elif args.model == 'LDTN':
            extractor, deepset = train_uf_y(tgt, utility_data[0], utility_data[1], train_set, test_set, args.epochs,
                                            src_weights=pretrain_net_file,
                                            out_dir=args.out_dir, logger=logger,
                                            betas=args.betas, weight_decay=args.weight_decay,
                                            model=args.model,
                                            ext_file=ext_file, ds_file=ds_file)
        else:
            extractor, deepset = train_uf(tgt, utility_data[0], train_set, test_set, args.epochs,
                                          pretrain_net_file=pretrain_net_file, out_dir=args.out_dir, logger=logger,
                                          model=args.model, ext_file=ext_file, ds_file=ds_file,
                                          lr_extractor=args.lr_extractor, lr_deepsets=args.lr_deepsets,
                                          device=args.device)
        extractor.eval()
        deepset.eval()

    ########################################### calculate Shapley Values  ###########################################
    shapley_value_uf = get_shapley(args.n_owners, data_owner_X, data_owner_y, extractor, deepset, valid_data)
    print(shapley_value_uf)
    print(np.argsort(shapley_value_uf)[::-1])
    random_uf = torch.rand(args.n_owners).numpy()
    accuracy_lines, scores = train_accuracy_line(shapley_value_uf, random_uf, data_owner_X, data_owner_y, proxy_model,
                                                 valid_data, logger)
    return accuracy_lines, scores
    # np.arange(args.n_owners)[::-1]


if __name__ == '__main__':
    args = experimental_setting()
    src = args.src
    tgt = args.tgt
    args.model = 'MPCLeNet' if tgt in ['mnist', 'usps'] else 'MPCDTN'
    args.num_cls = 10
    proxy_model = logistic_data_to_acc

    if src is None:
        csv_file = os.path.join(args.out_dir, 'effective/models/metric_{}_{}_{}.csv'.format(tgt, args.n_owners, args.data_size))
        args.out_dir = os.path.join(args.out_dir, 'effective/models/{}_share_{}/'.format(tgt, args.share))
        args.out_feature_dir = os.path.join(args.out_feature_dir,
                                            'effective/feature/{}_share_{}'.format(tgt, args.share))
        logfile = os.path.join(args.out_dir, 'output_{}_{}_{}.log'.format(tgt, args.n_owners, args.data_size))
        pretrain_net_file = os.path.join(args.out_dir, '{}_pretrain_{:s}.pth'.format(args.model, tgt))
        x_path = '{}/s1_x_{}_{}_{}.npy'.format(args.out_feature_dir, args.n_owners, args.data_size, tgt)
        y_path = '{}/s1_y_{}_{}_{}.npy'.format(args.out_feature_dir, args.n_owners, args.data_size, tgt)
        file_name = os.path.join(args.out_dir, 'output_{}_{}.xlsx'.format(args.n_owners, args.data_size))

    else:
        csv_file = os.path.join(args.out_dir,
                                'effective/models/metric_{}2{}_{}_{}.csv'.format(src, tgt, args.n_owners, args.data_size))
        args.out_dir = os.path.join(args.out_dir, 'effective/models/{}2{}_share_{}/'.format(src, tgt, args.share))
        args.out_feature_dir = os.path.join(args.out_feature_dir,
                                            'effective/feature/{}2{}_share_{}'.format(src, tgt, args.share))
        logfile = os.path.join(args.out_dir, 'output_{}2{}_{}_{}.log'.format(src, tgt, args.n_owners, args.data_size))
        pretrain_net_file = os.path.join(args.out_dir, '{}_pretrain_{:s}.pth'.format(args.model, src))
        x_path = '{}/s3_x_{}_{}_{}.npy'.format(args.out_feature_dir, args.n_owners, args.data_size, src)
        y_path = '{}/s3_y_{}_{}_{}.npy'.format(args.out_feature_dir, args.n_owners, args.data_size, src)
        file_name = os.path.join(args.out_dir, 'output_{}_{}.xlsx'.format(args.n_owners, args.data_size))
    # os.makedirs("../outputs/effective/models1/")
    if not os.path.exists(args.out_dir):
        os.makedirs(args.out_dir)
    if not os.path.exists(args.out_feature_dir):
        os.makedirs(args.out_feature_dir)
    if os.path.exists(logfile):
        os.remove(logfile)

    if os.path.exists(file_name):
        wb = load_workbook(file_name)
    else:
        wb = Workbook()
    wl2h = wb.create_sheet('{}_l2h'.format(tgt))
    wl2h_random = wb.create_sheet('{}_l2h_random'.format(tgt))
    wh2l = wb.create_sheet('{}_h2l'.format(tgt))
    wh2l_random = wb.create_sheet('{}_h2l_random'.format(tgt))
    ws_metric = wb.create_sheet('{}_metric'.format(tgt))
    progbar = tqdm(range(args.repeats), file=sys.stdout, position=0, leave=False)
    for i in progbar:
        seed_everything(args.seed)
        accuracy_lines, scores = main()
        args.seed += 1
        l2h, random_l2h, h2l, random_h2l = accuracy_lines
        scorel2h, scoreh2l = scores
        print(scorel2h, scoreh2l)

        with open(csv_file, 'a') as f:
            writer_object = writer(f)
            results = [args.share, scorel2h[0], scoreh2l[0]]
            writer_object.writerow(results)
        f.close()

        ws_metric.append([args.n_owners, args.data_size] + scorel2h + scoreh2l)
        wl2h.append(list([args.n_owners, args.data_size] + list(l2h)))
        wl2h_random.append(list([args.n_owners, args.data_size] + list(random_l2h)))
        wh2l.append(list([args.n_owners, args.data_size] + list(h2l)))
        wh2l_random.append(list([args.n_owners, args.data_size] + list(random_h2l)))
    wb.save(file_name)
